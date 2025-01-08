import os
from django.conf import settings
import matplotlib.pyplot
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from model.models import *
import numpy
from scipy.stats import norm
import io
from openpyxl.drawing.image import Image

#Predefines a light gray header for all the included tables.
table_header_color = PatternFill(start_color='E9E9E9', end_color='E9E9E9', fill_type='solid')
font_size = 13
text_font = Font(name='Times New Roman', size=font_size)
cell_alignment = Alignment(horizontal='center')

#Global colors for cells.
purple_fill = PatternFill(start_color="785EF0", end_color="785EF0", fill_type="solid")
blue_fill = PatternFill(start_color="648FFF", end_color="648FFF", fill_type="solid")
green_fill = PatternFill(start_color="009E73", end_color="009E73", fill_type="solid")
yellow_fill = PatternFill(start_color="FFB000", end_color="FFB000", fill_type="solid")
orange_fill = PatternFill(start_color="FE6100", end_color="FE6100", fill_type="solid")
pink_fill = PatternFill(start_color="DC267F", end_color="DC267F", fill_type="solid")


def generateSpreadsheet(patient: str):
    matplotlib.use('agg')
    link = doctorLink.objects.get(linkID=patient)

    indv_tab_df = pd.DataFrame(list(individualTable.objects.filter(
        linkID = link
    ).values()))

    indv_res_df = pd.DataFrame(list(individualResultTable.objects.filter(
        indTable__linkID = link
    ).values()))

    patient_age = indv_tab_df['age'].unique()
    age_agg_data = []
    for age in patient_age:
        pat_agg_data = getAggregateDataForAgeSummary(age)
        age_agg_data.append(pat_agg_data)
    agg_df = pd.DataFrame(age_agg_data)

    latency_df = pd.DataFrame(list(latency.objects.filter(
        indResTable__in = indv_res_df['id']
    ).values()))

    #Merges the Indivdual Results tabel and the Latency Table.
    full_df = pd.merge(indv_res_df, latency_df, left_on='id', right_on='indResTable_id', how='left')

    #Adds the Latencies to the middle of the table.
    indRes_index = full_df.columns.get_loc('patientAnswerString')
    latency_columns = [col for col in full_df.columns if col.startswith('latency')]
    final_columns = (list(full_df.columns[:indRes_index + 1]) + 
                 latency_columns + 
                 list(full_df.columns[indRes_index + 1:].difference(latency_columns)))
    
    #Deletes the unneeded columns.
    full_res_df = full_df[final_columns]
    full_res_df = full_res_df.drop(columns=['id_x', 'indTable_id', 'id_y', 'indResTable_id'])
    indv_tab_df = indv_tab_df.drop(columns=['linkID_id'])

    full_res_df = full_res_df[[col for col in full_res_df.columns if col != 'endingDayAndTime'] + ['endingDayAndTime']]

    #Defined columns for the Indivudal Table rounding.
    indv_decimal_columns = ['avgLatency1', 'avgLatency2', 'avgLatency3', 
                            'avgLatency4', 'avgLatency5', 'avgDuration', 
                            'avg4SpanScore', 'avg5SpanScore', 
                            'avgScoreDigit', 'avgScoreMixed', 'avgScore']
    
    #For loop to round the columns.
    for col in indv_decimal_columns:
        if col in indv_tab_df.columns:
            indv_tab_df[col] = indv_tab_df[col].round(2)

    #Defined columns for the table rounding.
    agg_decomal_columns = ['avgLatency1', 'sdLatency1', 'avgLatency2',
                           'sdLatency2', 'avgLatency3', 'sdLatency3',
                           'avgLatency4', 'sdLatency4', 'avgLatency5',
                           'sdLatency5', 'avgDuration', 'sdDuration',
                           'avg4SpanScore', 'sd4SpanScore', 'avg5SpanScore',
                           'sd5SpanScore', 'avgScoreDigit', 'sdScoreDigit',
                           'avgScoreMixed', 'sdScoreMixed', 'avgScore',
                           'sdScore']

    #Defined columns for the Indivdual Results table that needs to be rounded.
    indvr_decimal_column = ['latency1', 'latency2', 'latency3', 
                            'latency4','latency5', 'duration']

    #Rounds the columns of the Indivdual Results Table.
    for col in indvr_decimal_column:
        if col in full_res_df.columns:
            full_res_df[col] = full_res_df[col].round(2)

    #Renames the columns of the Indivdual Table .
    indv_tab_df.columns = ['Patient Test ID', 'Time Started', 'Age',
                           'Test Status', 'Average Latency 1', 'Average Latency 2',
                           'Average Latency 3', 'Average Latency 4', 'Average Latency 5',
                           'Average Test Duration', 'Average 4 Span Score', 'Average 5 Span Score',
                           'Average Score Digit', 'Average Score Mixed', 'Average Score']
    #Renames the columns of the Aggergate Datatable.
    agg_df.columns = ['Minimum Age', 'Maximum Age', 'Tests for Age Range',
                        'Average Latency 1', 'Latency 1 Standard Deviation', 'Average Latency 2',
                        'Latency 2 Standard Deviation', 'Average Latency 3', 'Latency 3 Standard Deviation',
                        'Average Latency 4', 'Latency 4 Standard Deviation', 'Average Latency 5',
                        'Latency 5 Standard Deviation', 'Average Duration', 'Duration Standard Deviation',
                        'Average 4 Span Score', '4 Span Standard Deviation', 'Average 5 Span Score',
                        '5 Span Score Standard Deviation', 'Average Score Digit', 'Score Digit Standard Deviation',
                        'Average Score Mixed', 'Mixed Score Standard Deviation', 'Average Score',
                        'Score Standard Deviation']
    
    stimuli_data = stimuli.objects.all().values('correctString')
    stimuli_df = pd.DataFrame(stimuli_data)
    correct_string_column = stimuli_df['correctString']
    full_res_df.insert(1, 'Correct Answer', correct_string_column)

    #Renames all the columns of the full Indivdual Results Table.
    full_res_df.columns = ['Test Type', 'Correct Answer', 'Patient Answer', 'Latency 1',
                           'Latency 2', 'Latency 3', 'Latency 4',
                           'Latency 5', 'Duration', 'Score',
                           'Start Time', 'End Time']
    
    stimuli_data = stimuli.objects.all().values('name')
    stimuli_df = pd.DataFrame(stimuli_data)
    full_res_df['Test Type'] = stimuli_df['name']

    #Get the path to the server temp folder where the spreadsheet will be saved.
    filePath = os.path.join(settings.BASE_DIR, 'temp/'+ patient + '.xlsx')

    #Function to extract the data to the spreadsheet.
    with pd.ExcelWriter(filePath, engine='openpyxl',) as writer:
        full_res_df.to_excel(writer,  index=False, sheet_name='Raw Data', startrow=0)
        full_res_df.to_excel(writer, index=False, sheet_name='Raw Data 2', startrow=0)

    workbook = load_workbook(filePath)
    sheet = workbook['Raw Data']
    sheet2 = workbook['Raw Data 2']
        
    for sheet in [sheet, sheet2]:
        for cell in sheet[1]: 
            cell.fill = table_header_color

    color_columns = ['Latency 1', 'Latency 2', 'Latency 3', 'Latency 4', 'Latency 5', 'Duration', 'Score']
    sd_columns = ['Latency 1 Standard Deviation', 'Latency 2 Standard Deviation', 'Latency 3 Standard Deviation', 'Latency 4 Standard Deviation', 'Latency 5 Standard Deviation',
                          'Duration Standard Deviation', 'Score Standard Deviation']
    mean_columns = ['Average Latency 1','Average Latency 2','Average Latency 3', 'Average Latency 4',
                    'Average Latency 5', 'Average Duration', 'Average Score' ]

    for i, col in enumerate(color_columns, start=4):
        avg = agg_df[mean_columns[i - 4]].iloc[0]
        sd = agg_df[sd_columns[i - 4]].iloc[0]

        for row in range(2, len(full_res_df) + 2):
            data_point = full_res_df.loc[row - 2, col]
            zscore_value, color = determineStdColor(data_point, avg, sd)
            
            cell = sheet2.cell(row=row, column=i)
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    

    sheet2['A30'] = '4 Span Score'
    sheet2['B30'] = '5 Span Score'
    sheet2['C30'] = 'Difference'
    sheet2['A33'] = '4 Digit Score'
    sheet2['B33'] = '5 Digit Score'
    sheet2['C33'] = 'Difference'
    sheet2['A36'] = '4 Mixed Score'
    sheet2['B36'] = '5 Mixed Score'
    sheet2['C36'] = 'Difference'
    

    header_font = Font(name='Times New Roman', size=13, bold=True)
    header_fill = PatternFill(start_color='E9E9E9', end_color='E9E9E9', fill_type='solid')


    for cell in ['A30', 'B30', 'C30', 'A33', 'B33', 'C33', 'A36', 'B36', 'C36']:
        sheet2[cell].font = header_font
        sheet2[cell].fill = header_fill

    custom_font = Font(name='Arial', size=14, bold=True)


    data = individualTable.objects.get(linkID=link).calculateExtraData()

    for idx, row in indv_tab_df.iterrows():
        score_4_span = round(row['Average 4 Span Score'],2)
        score_5_span = round(row['Average 5 Span Score'],2)
        difference = score_5_span - score_4_span
        score_4_span = round(data['4SpanDigitScore'],2)
        score_digit = round(data['5SpanDigitScore'],2)
        score_4_digit_value = score_digit - score_4_span
        score_4_span = round(data['4SpanMixedScore'],2)
        score_mixed = round(data['5SpanMixedScore'],2)
        score_4_mixed_value = score_mixed - score_4_span

    score_4_span_cell = sheet2.cell(row=31, column=1, value=score_4_span).font = custom_font
    score_5_span_cell = sheet2.cell(row=31, column=2, value=score_5_span).font = custom_font
    difference_cell = sheet2.cell(row=31, column=3, value=difference)
    score_4_span_cell = sheet2.cell(row=34, column=1, value=score_4_span).font = custom_font
    score_digit_cell = sheet2.cell(row=34, column=2, value=score_digit).font = custom_font
    score_4_digit_value_cell = sheet2.cell(row=34, column=3, value=score_4_digit_value)
    score_4_span_cell = sheet2.cell(row=37, column=1, value=score_4_span).font = custom_font
    score_mixed_cell = sheet2.cell(row=37, column=2, value=score_mixed).font = custom_font
    score_4_mixed_value_cell = sheet2.cell(row=37, column=3, value=score_4_mixed_value)

    
    difference_cell.font = custom_font
    score_digit_cell.font = custom_font
    score_4_digit_value_cell.font = custom_font
    score_mixed_cell.font = custom_font
    score_4_mixed_value_cell.font = custom_font

    if difference > 0:
        difference_cell.fill = green_fill
    elif difference < 0:
        difference_cell.fill = orange_fill

    if score_4_digit_value > 0:
        score_4_digit_value_cell.fill = green_fill
    elif score_4_digit_value < 0:
        score_4_digit_value_cell.fill = orange_fill

    if score_4_mixed_value > 0:
        score_4_mixed_value_cell.fill = green_fill
    elif score_4_mixed_value < 0:
        score_4_mixed_value_cell.fill = orange_fill

    

    column_widths = {'A': 15, 'B': 15, 'C': 15, 'E': 15, 'F': 15, 'G': 15, 'I': 15, 'J': 15, 'K': 15,
    'M': 15, 'N': 15, 'O': 15, 'Q': 15, 'R': 15, 'S': 15, 'U': 15, 'V': 15, 'W': 15}
    for column, width in column_widths.items():
        sheet2.column_dimensions[column].width = width


    #Adds legend to sheet2.
    sheet2.cell(row=22, column=2, value="Standard Deviation Legend:").fill = table_header_color
    sheet2.cell(row=23, column=2, value=">= 2 Standard Deviations").fill = purple_fill
    sheet2.cell(row=24, column=2, value=">= 1 Standard Deviations").fill = blue_fill
    sheet2.cell(row=25, column=2, value=">= 0 Standard Deviations").fill = green_fill
    sheet2.cell(row=26, column=2, value=">= -1 Standard Deviations").fill = yellow_fill
    sheet2.cell(row=27, column=2, value=">= -2 Standard Deviations").fill = orange_fill
    sheet2.cell(row=28, column=2, value="< -2 Standard Deviations").fill = pink_fill
    createLegend(sheet=sheet2,start_row=22,start_col=2)

    sheet2.cell(row=39, column=2, value="Span Comparison Legend")
    sheet2.cell(row=40, column=2, value="Positive Difference").fill = green_fill
    sheet2.cell(row=41, column=2, value="Negative Difference").fill = orange_fill
    

    sheet_names = ['Raw Data', 'Raw Data 2']


    #Adds font and cell alignment to the feels of the Raw Data sheet.
    for sheet_name in sheet_names:
        sheet = workbook[sheet_name]
        for row in range(1, sheet.max_row + 1):
            for col in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=row, column=col)
                if cell.value is not None:
                    cell.font = text_font
                    cell.alignment = cell_alignment
        #Auto-fit the spreadsheet for the max length of the cell.
        #+10 padding added for some cells not correctly adjusting.
        columns = sheet.max_column  
        for col in range(1, columns + 1):  
            max_length = 0
            for row_index in [1, 5, 8, 23]:  
                cell_value = sheet.cell(row=row_index, column=col).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            adjusted_width = max_length + 10  
            sheet.column_dimensions[sheet.cell(row=1, column=col).column_letter].width = adjusted_width

    # Summary Aggregate Table
    indTable = individualTable.objects.get(linkID=link)
    indData = indTable.getAllFields()
    extraData = indTable.calculateExtraData()
    patientData = {**indData, **extraData}
    age = int(indv_tab_df['Age'].iloc[0])
    aggregateData = getAggregateDataForAgeComplete(age)
    zscores = getAllZScore(patientData, age)
    aggregate_sheet = workbook.create_sheet(title="Aggregate Info")
    timeHeader = ['Latency 1','Latency 2','Latency 3','Latency 4','Latency 5','Duration Per Test','Duration Whole Test']
    timeFields = ['avgLatency1','avgLatency2','avgLatency3','avgLatency4','avgLatency5','avgDuration','DurationWholeTest']
    scoreHeader = ['4 Span Digit','4 Span Mixed','5 Span Digit','5 Span Mixed','4 Span','5 Span','Digit','Mixed','Overall']
    serialScores = ['4SpanDigitScore','4SpanMixedScore','5SpanDigitScore','5SpanMixedScore','avg4SpanScore','avg5SpanScore','avgScoreDigit','avgScoreMixed','avgScore']
    unorderedScores = ['4SpanDigitScoreUnordered','4SpanMixedScoreUnordered','5SpanDigitScoreUnordered','5SpanMixedScoreUnordered','4SpanScoreUnordered','5SpanScoreUnordered','ScoreDigitUnordered','ScoreMixedUnordered','ScoreUnordered']
    stimHeaders = ['Position 1','Position 2','Position 3','Position 4','Position 5']
    stimScores = ['ScoreStim1','ScoreStim2','ScoreStim3','ScoreStim4','ScoreStim5']
    titleTime = aggregate_sheet.cell(row=1,column=1,value="Time")
    titleTime.font = Font(name='Times New Roman', size=18, bold=True)
    titleTime.alignment = Alignment(horizontal='center')
    createAggregateTable(aggregate_sheet=aggregate_sheet,headers=timeHeader,fields=timeFields,patientData=patientData,aggregateData=aggregateData,zscores=zscores,start_row=2)
    titleSerial = aggregate_sheet.cell(row=9,column=1,value="Serial Score")
    titleSerial.font = Font(name='Times New Roman', size=18, bold=True)
    titleSerial.alignment = Alignment(horizontal='center')
    createAggregateTable(aggregate_sheet=aggregate_sheet,headers=scoreHeader,fields=serialScores,patientData=patientData,aggregateData=aggregateData,zscores=zscores,start_row=10)
    titleUnordered = aggregate_sheet.cell(row=17,column=1,value="Unordered Score")
    titleUnordered.font = Font(name='Times New Roman', size=18, bold=True)
    titleUnordered.alignment = Alignment(horizontal='center')
    createAggregateTable(aggregate_sheet=aggregate_sheet,headers=scoreHeader,fields=unorderedScores,patientData=patientData,aggregateData=aggregateData,zscores=zscores,start_row=18)
    titleStim = aggregate_sheet.cell(row=25,column=1,value="Response Position Score")
    titleStim.font = Font(name='Times New Roman', size=18, bold=True)
    titleStim.alignment = Alignment(horizontal='center')
    createAggregateTable(aggregate_sheet=aggregate_sheet,headers=stimHeaders,fields=stimScores,patientData=patientData,aggregateData=aggregateData,zscores=zscores,start_row=26)
    # Format the cell sizes
    columns = aggregate_sheet.max_column 
    for col in range(1, columns + 1):  
        max_length = 0
        for row_index in [1,2,9,10,17,18,25,26]:  
            cell_value = aggregate_sheet.cell(row=row_index, column=col).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        adjusted_width = max_length + 10  
        aggregate_sheet.column_dimensions[aggregate_sheet.cell(row=1, column=col).column_letter].width = adjusted_width
    createLegend(aggregate_sheet,33,1)

    # Graph Sheets
    time_graph_sheet = workbook.create_sheet(title="Time Graphs")
    start_row = 1
    start_col = 1
    position = 'A'
    for i in range(len(timeFields)):
        data_point = round(patientData[timeFields[i]],2)
        field = timeFields[i]
        if not field.startswith('avg'):
            field = 'avg' + field
        avg = round(aggregateData[field],2)
        field = timeFields[i]
        if field.startswith('avg'):
            field = field[3:]
        field = 'sd' + field
        sd = round(aggregateData[field],2)
        createGraphTable(time_graph_sheet,timeHeader[i],data_point,avg,sd,start_row,start_col)
        pos = position + str((start_row+5))
        createGraph(field=field,data_point=data_point,avg=avg,sd=sd,graph_sheet=time_graph_sheet,position=pos)
        if i % 2 == 0:
            start_col += 5
            position = 'F'
        else:
            start_row += 20
            start_col -= 5
            position = 'A'
    createLegend(time_graph_sheet,1,10)
    # Format the cell sizes
    columns = time_graph_sheet.max_column 
    for col in range(1, columns + 1):  
        max_length = 0
        for row_index in [1,2,3,4,5,21,22,23,24,25,41,42,43,44,45,61,62,63,64,65,81,82,83,84,85]:  
            cell_value = time_graph_sheet.cell(row=row_index, column=col).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        adjusted_width = max_length + 10  
        time_graph_sheet.column_dimensions[time_graph_sheet.cell(row=1, column=col).column_letter].width = adjusted_width
    serial_graph_sheet = workbook.create_sheet(title="Serial Score Graphs")
    start_row = 1
    start_col = 1
    position = 'A'
    for i in range(len(serialScores)):
        data_point = round(patientData[serialScores[i]],2)
        field = serialScores[i]
        if not field.startswith('avg'):
            field = 'avg' + field
        avg = round(aggregateData[field],2)
        field = serialScores[i]
        if field.startswith('avg'):
            field = field[3:]
        field = 'sd' + field
        sd = round(aggregateData[field],2)
        createGraphTable(serial_graph_sheet,scoreHeader[i],data_point,avg,sd,start_row,start_col)
        pos = position + str((start_row+5))
        createGraph(field=field,data_point=data_point,avg=avg,sd=sd,graph_sheet=serial_graph_sheet,position=pos)
        if i % 2 == 0:
            start_col += 5
            position = 'F'
        else:
            start_row += 20
            start_col -= 5
            position = 'A'
    createLegend(serial_graph_sheet,1,10)
    # Format the cell sizes
    columns = serial_graph_sheet.max_column 
    for col in range(1, columns + 1):  
        max_length = 0
        for row_index in [1,2,3,4,5,21,22,23,24,25,41,42,43,44,45,61,62,63,64,65,81,82,83,84,85]:  
            cell_value = serial_graph_sheet.cell(row=row_index, column=col).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        adjusted_width = max_length + 10  
        serial_graph_sheet.column_dimensions[serial_graph_sheet.cell(row=1, column=col).column_letter].width = adjusted_width
    unordered_graph_sheet = workbook.create_sheet(title="Unordered Score Graphs")
    start_row = 1
    start_col = 1
    position = 'A'
    for i in range(len(unorderedScores)):
        data_point = round(patientData[unorderedScores[i]],2)
        field = unorderedScores[i]
        if not field.startswith('avg'):
            field = 'avg' + field
        avg = round(aggregateData[field],2)
        field = unorderedScores[i]
        if field.startswith('avg'):
            field = field[3:]
        field = 'sd' + field
        sd = round(aggregateData[field],2)
        createGraphTable(unordered_graph_sheet,scoreHeader[i],data_point,avg,sd,start_row,start_col)
        pos = position + str((start_row+5))
        createGraph(field=field,data_point=data_point,avg=avg,sd=sd,graph_sheet=unordered_graph_sheet,position=pos)
        if i % 2 == 0:
            start_col += 5
            position = 'F'
        else:
            start_row += 20
            start_col -= 5
            position = 'A'
    createLegend(unordered_graph_sheet,1,10)
    # Format the cell sizes
    columns = unordered_graph_sheet.max_column 
    for col in range(1, columns + 1):  
        max_length = 0
        for row_index in [1,2,3,4,5,21,22,23,24,25,41,42,43,44,45,61,62,63,64,65,81,82,83,84,85]:  
            cell_value = unordered_graph_sheet.cell(row=row_index, column=col).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        adjusted_width = max_length + 10  
        unordered_graph_sheet.column_dimensions[unordered_graph_sheet.cell(row=1, column=col).column_letter].width = adjusted_width
    stim_graph_sheet = workbook.create_sheet(title="Response Position Score Graphs")
    start_row = 1
    start_col = 1
    position = 'A'
    for i in range(len(stimScores)):
        data_point = round(patientData[stimScores[i]],2)
        field = stimScores[i]
        if not field.startswith('avg'):
            field = 'avg' + field
        avg = round(aggregateData[field],2)
        field = stimScores[i]
        if field.startswith('avg'):
            field = field[3:]
        field = 'sd' + field
        sd = round(aggregateData[field],2)
        createGraphTable(stim_graph_sheet,stimHeaders[i],data_point,avg,sd,start_row,start_col)
        pos = position + str((start_row+5))
        createGraph(field=field,data_point=data_point,avg=avg,sd=sd,graph_sheet=stim_graph_sheet,position=pos)
        if i % 2 == 0:
            start_col += 5
            position = 'F'
        else:
            start_row += 20
            start_col -= 5
            position = 'A'
    createLegend(stim_graph_sheet,1,10)
    # Format the cell sizes
    columns = stim_graph_sheet.max_column 
    for col in range(1, columns + 1):  
        max_length = 0
        for row_index in [1,2,3,4,5,21,22,23,24,25,41,42,43,44,45,61,62,63,64,65,81,82,83,84,85]:  
            cell_value = stim_graph_sheet.cell(row=row_index, column=col).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        adjusted_width = max_length + 10  
        stim_graph_sheet.column_dimensions[stim_graph_sheet.cell(row=1, column=col).column_letter].width = adjusted_width

    #Saves the workbook to the Users download folder.
    workbook.save(filePath)           

    
def determineStdColor(data_point, avg, sd):
    zscore = (data_point - avg) / sd
    if zscore >= 2:
        # Purple
        color = '785EF0'
    elif zscore >= 1:
        # Blue
        color = '648FFF'
    elif zscore >= 0:
        # Green
        color = '009E73'
    elif zscore >= -1:
        # Yellow
        color = 'FFB000'
    elif zscore >= -2:
        # Orange
        color = 'FE6100'
    else:
        # Pinkish Red
        color = 'DC267F'
    return zscore, color

    

def createGraph(field, data_point, avg, sd, graph_sheet, position):
    # Feel like the nums should be based off the data and not just avg
    nums = numpy.linspace(avg - 3 *sd, avg + 3 * sd, 1000)
    # data_points = numpy.array([entry[field] for entry in individualTable.objects.filter(minAge__lte=age, maxAge__gte=age).values(field)])
    # data_points = individualTable.objects.filter(age__gt=aggregated_data_row.minAge-1, age__lt=aggregated_data_row.maxAge+1, completed=individualTable.completedStatus.COMPLETED)
    # nums = numpy.sort(data_points)
    distribution = norm.pdf(nums, avg, sd)
    fig, ax = matplotlib.pyplot.subplots()
    ax.plot(nums, distribution, label=f"{field} Normal Distribution", color='blue')
    ax.fill_between(nums, distribution, alpha=0.5)
    ax.set_title(f"{field} Normal Distribution")
    ax.set_xlabel(f"{field}")
    ax.set_ylabel('Likelihood')
    ax.axvline(x=data_point, color='red',linestyle='--',label=f'Data Point: {data_point}')
    buffer = io.BytesIO()
    fig.savefig(buffer, format='png')
    buffer.seek(0)
    img = Image(buffer)
    img.width = int(img.width * 0.5)
    img.height = int(img.height * 0.5) 
    graph_sheet.add_image(img, position)
    matplotlib.pyplot.close(fig)


def createAggregateTable(aggregate_sheet,headers,fields,patientData,aggregateData,zscores,start_row):
    table_header_color = PatternFill(start_color='E9E9E9', end_color='E9E9E9', fill_type='solid')
    table_header_font = Font(name='Times New Roman', size=13, bold=True)
    table_font = Font(name='Times New Roman', size=13)
    table_alignment = Alignment(horizontal='center')
    # Create Row Labels
    cell = aggregate_sheet.cell(row=start_row,column=1,value="Field")
    cell.font = table_header_font
    cell.alignment = table_alignment
    cell.fill = table_header_color
    cell = aggregate_sheet.cell(row=start_row+1,column=1,value="Patient")
    cell.font = table_header_font
    cell.alignment = table_alignment
    cell.fill = table_header_color
    cell = aggregate_sheet.cell(row=start_row+2,column=1,value="Aggregate")
    cell.font = table_header_font
    cell.alignment = table_alignment
    cell.fill = table_header_color
    cell = aggregate_sheet.cell(row=start_row+3,column=1,value="Std")
    cell.font = table_header_font
    cell.alignment = table_alignment
    cell.fill = table_header_color
    cell = aggregate_sheet.cell(row=start_row+4,column=1,value="Z-Score")
    cell.font = table_header_font
    cell.alignment = table_alignment
    cell.fill = table_header_color
    # Fill table columns
    for i in range(len(headers)):
        # Header
        cell = aggregate_sheet.cell(row=start_row,column=i+2,value=f"{headers[i]}")
        cell.font = table_header_font
        cell.alignment = table_alignment
        cell.fill = table_header_color
        # Patient Data Point
        data_point = round(patientData[fields[i]],2)
        cell = aggregate_sheet.cell(row=start_row+1,column=i+2,value=f"{data_point}")
        cell.font = table_font
        cell.alignment = table_alignment
        # Aggregate Average
        field = fields[i]
        if not field.startswith('avg'):
            field = 'avg' + field
        avg = round(aggregateData[field],2)
        cell = aggregate_sheet.cell(row=start_row+2,column=i+2,value=f"{avg}")
        cell.font = table_font
        cell.alignment = table_alignment
        # Aggregate Standard Deviation
        field = fields[i]
        if field.startswith('avg'):
            field = field[3:]
        field = 'sd' + field
        sd = round(aggregateData[field],2)
        cell = aggregate_sheet.cell(row=start_row+3,column=i+2,value=f"{sd}")
        cell.font = table_font
        cell.alignment = table_alignment
        # Z-Score
        field = fields[i]
        if field.startswith('avg'):
            field = field[3:]
        zscore = round(zscores[field],2)
        cell = aggregate_sheet.cell(row=start_row+4,column=i+2,value=f"{zscore}")
        cell.font = table_font
        cell.alignment = table_alignment
        _, color = determineStdColor(data_point,avg,sd)
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')

def createLegend(sheet, start_row, start_col):
    sheet.cell(row=start_row, column=start_col, value="Standard Deviation Legend:")
    sheet.cell(row=start_row+1, column=start_col, value=">= 2 Standard Deviations").fill = purple_fill
    sheet.cell(row=start_row+2, column=start_col, value=">= 1 Standard Deviations").fill = blue_fill
    sheet.cell(row=start_row+3, column=start_col, value=">= 0 Standard Deviations").fill = green_fill
    sheet.cell(row=start_row+4, column=start_col, value=">= -1 Standard Deviations").fill = yellow_fill
    sheet.cell(row=start_row+5, column=start_col, value=">= -2 Standard Deviations").fill = orange_fill
    sheet.cell(row=start_row+6, column=start_col, value="< -2 Standard Deviations").fill = pink_fill

def createGraphTable(graph_sheet, field, data_point, avg, sd, start_row, start_col):
    table_header_color = PatternFill(start_color='E9E9E9', end_color='E9E9E9', fill_type='solid')
    table_header_font = Font(name='Times New Roman', size=13, bold=True)
    table_font = Font(name='Times New Roman', size=13)
    table_alignment = Alignment(horizontal='center')
    # Left column labels
    cell = graph_sheet.cell(row=start_row,column=start_col,value="Field")
    cell.font = table_header_font
    cell.alignment = table_alignment
    cell.fill = table_header_color
    cell = graph_sheet.cell(row=start_row+1,column=start_col,value="Patient")
    cell.font = table_header_font
    cell.alignment = table_alignment
    cell.fill = table_header_color
    cell = graph_sheet.cell(row=start_row+2,column=start_col,value="Average")
    cell.font = table_header_font
    cell.alignment = table_alignment
    cell.fill = table_header_color
    cell = graph_sheet.cell(row=start_row+3,column=start_col,value="Standard Deviation")
    cell.font = table_header_font
    cell.alignment = table_alignment
    cell.fill = table_header_color
    cell = graph_sheet.cell(row=start_row+4,column=start_col,value="Z-Score")
    cell.font = table_header_font
    cell.alignment = table_alignment
    cell.fill = table_header_color
    cell = graph_sheet.cell(row=start_row,column=start_col+1,value=f"{field}")
    cell.font = table_header_font
    cell.alignment = table_alignment
    cell.fill = table_header_color
    cell = graph_sheet.cell(row=start_row+1,column=start_col+1,value=f"{data_point}")
    cell.font = table_font
    cell.alignment = table_alignment
    cell = graph_sheet.cell(row=start_row+2,column=start_col+1,value=f"{avg}")
    cell.font = table_font
    cell.alignment = table_alignment
    cell = graph_sheet.cell(row=start_row+3,column=start_col+1,value=f"{sd}")
    cell.font = table_font
    cell.alignment = table_alignment
    zscore, color = determineStdColor(data_point,avg,sd)
    cell = graph_sheet.cell(row=start_row+4,column=start_col+1,value=f"{round(zscore,2)}")
    cell.font = table_font
    cell.alignment = table_alignment
    cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
    
